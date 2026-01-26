from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer, Order, Receiver
from .forms import NameLookupForm, CustomerForm, OrderForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl import load_workbook


def entry(request):
    return render(request, "entry.html", {
        "is_staff": request.user.is_authenticated and request.user.is_staff
    })

def order_start(request):
    matches = []
    query_name = ""

    if request.method == "POST":
        # 신규 고객 생성 폼 제출
        if request.POST.get("create_new") == "1":
            cform = CustomerForm(request.POST)
            if cform.is_valid():
                customer = cform.save()
                request.session["customer_id"] = customer.id
                return redirect("order_new")  # 다음 단계에서 만들 예정
        else:
            cform = CustomerForm()

        # 이름 조회
        lookup = NameLookupForm(request.POST)
        if lookup.is_valid():
            query_name = lookup.cleaned_data["name"].strip()
            matches = Customer.objects.filter(name__icontains=query_name)[:10]

        return render(request, "order_start.html", {
            "lookup": lookup,
            "matches": matches,
            "cform": cform,
            "query_name": query_name,
        })

    return render(request, "order_start.html", {
        "lookup": NameLookupForm(),
        "matches": [],
        "cform": CustomerForm(),
        "query_name": "",
    })
def order_name(request):
    if request.method == "POST":
        form = NameLookupForm(request.POST)
        if form.is_valid():
            request.session["lookup_name"] = form.cleaned_data["name"].strip()
            return redirect("order_customer")
    else:
        form = NameLookupForm()
    return render(request, "order_name.html", {"form": form})

from django.shortcuts import render, redirect, get_object_or_404

def order_customer(request):
    name = request.session.get("lookup_name", "").strip()
    if not name:
        return redirect("order_name")

    matches = Customer.objects.filter(name__icontains=name).order_by("name")[:10]

    # 기존 고객 선택
    if request.method == "POST" and request.POST.get("selected_customer_id"):
        cid = int(request.POST["selected_customer_id"])
        customer = get_object_or_404(Customer, id=cid)
        request.session["customer_id"] = customer.id
        return redirect("order_new")  # 다음 단계에서 만들 주문 화면

    # 신규 고객 생성
    if request.method == "POST" and request.POST.get("create_new") == "1":
        cform = CustomerForm(request.POST)
        if cform.is_valid():
            customer = cform.save()
            request.session["customer_id"] = customer.id
            return redirect("order_new")
    else:
        cform = CustomerForm(initial={"name": name})

    return render(request, "order_customer.html", {
        "name": name,
        "matches": matches,
        "cform": cform,
    })

def order_new(request):
    customer_id = request.session.get("customer_id")
    if not customer_id:
        return redirect("order_name")  # 고객 선택부터 다시

    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == "POST":
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.customer = customer
            order.save()  # total_price 자동 계산됨
            request.session["last_order_id"] = order.id
            return redirect("order_done")
    else:
        form = OrderForm(initial={
            "receiver_name": customer.name,
            "receiver_phone": customer.phone,
            "receiver_address": customer.address,
        })

    return render(request, "order_form.html", {"form": form, "customer": customer})

def order_done(request):
    order_id = request.session.get("last_order_id")
    customer_id = request.session.get("customer_id")
    if not order_id or not customer_id:
        return redirect("order_name")

    customer = get_object_or_404(Customer, id=customer_id)
    order = get_object_or_404(Order, id=order_id, customer=customer)
    return render(request, "order_done.html", {"customer": customer, "order": order})

def my_orders(request):
    customer_id = request.session.get("customer_id")
    if not customer_id:
        return redirect("order_name")

    customer = get_object_or_404(Customer, id=customer_id)
    orders = Order.objects.filter(customer=customer).order_by("-created_at")

    return render(request, "my_orders.html", {
        "customer": customer,
        "orders": orders,
    })

@login_required
def export_customers_xlsx(request):
    if not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    ws.append(["customer_name","customer_phone","customer_address","receiver_name","receiver_phone","receiver_address"])

    customers = Customer.objects.order_by("name", "phone")
    for c in customers:
        receivers = Receiver.objects.filter(customer=c).order_by("name", "phone")
        if receivers.exists():
            for r in receivers:
                ws.append([c.name, c.phone, c.address, r.name, r.phone, r.address])
        else:
            ws.append([c.name, c.phone, c.address, "", "", ""])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="customers.xlsx"'
    wb.save(resp)
    return resp

def export_orders_xlsx(request):
    if not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append([
        "created_at",
        "buyer_name", "buyer_phone",
        "receiver_name", "receiver_phone", "receiver_address",
        "product", "quantity", "unit_price", "total_price",
    ])

    for o in Order.objects.select_related("customer").order_by("-created_at"):
        ws.append([
            o.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            o.customer.name,
            o.customer.phone,
            o.receiver_name,
            o.receiver_phone,
            o.receiver_address,
            o.get_product_display(),
            o.quantity,
            o.unit_price,
            o.total_price,
        ])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
    wb.save(resp)
    return resp


def import_customers_xlsx(request):
    if not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)

    if request.method == "POST" and request.FILES.get("file"):
        wb = load_workbook(request.FILES["file"])
        ws = wb.active

        # 1) 헤더 읽기
        headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(headers)}

        required = ["customer_name", "customer_phone", "customer_address", "receiver_name", "receiver_phone", "receiver_address"]
        missing = [c for c in required if c not in idx]
        if missing:
            return HttpResponse(f"Missing columns: {missing}", status=400)

        # 2) 데이터 처리
        for row in ws.iter_rows(min_row=2, values_only=True):
            c_name = (row[idx["customer_name"]] or "").strip()
            c_phone = (row[idx["customer_phone"]] or "").strip()
            c_addr = (row[idx["customer_address"]] or "").strip()

            r_name = (row[idx["receiver_name"]] or "").strip()
            r_phone = (row[idx["receiver_phone"]] or "").strip()
            r_addr = (row[idx["receiver_address"]] or "").strip()

            if not c_name or not c_phone:
                continue  # 고객 식별 불가하면 스킵

            customer, _ = Customer.objects.update_or_create(
                name=c_name,
                phone=c_phone,
                defaults={"address": c_addr},
            )

            # 받는사람 정보가 없으면(빈 줄) receiver는 생성하지 않음
            if r_name or r_phone or r_addr:
                Receiver.objects.update_or_create(
                    customer=customer,
                    name=r_name or c_name,       # 받는사람 이름이 비면 고객 이름으로
                    phone=r_phone or c_phone,     # 받는사람 전화가 비면 고객 전화로
                    defaults={"address": r_addr or c_addr},
                )

        return redirect("/admin/")

    return render(request, "import_customers.html")